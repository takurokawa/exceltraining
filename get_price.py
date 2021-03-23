import openpyxl
import re

def value_get(base_sheet, search_sheet, id, search_str):
    for row_data in base_sheet.iter_rows(min_row=4):
        if str(row_data[2].value) == id:
            base_value = row_data[43].value
            break

    for row_data in search_sheet.iter_rows(min_row=2):
        for cell in row_data:
            if cell.value == None:
                continue
            elif re.search(r'2021-01-01',str(cell.value)):
                start_col = cell.column
            elif re.search(search_str, str(cell.value)):
                end_col = cell.column
                break
        break

    for row_data in search_sheet.iter_rows(min_row=3):
        if str(row_data[1].value) == id:
            flag = False
            dec = 0
            for cell in row_data:

                if cell.column == start_col:
                    flag = True
                if flag == True:
                    dec += int(cell.value)
                    if cell.column == end_col:
                        break

    result = int(base_value) + dec
    return result


def main():
    target_file = 'FILENAME.xlsx'
    base_value_sheet = 'BASE'
    decrement_value_sheet  = 'SEARCH'

    ids = input('Input searching values:').split()
    ids.sort()

    while True:
        date_data = input('Input year/month:')
        if re.fullmatch(r'202[0-9]/([1-9]|10|11|12)',date_data):
            date_list = date_data.split(sep='/')
            date_list[1] = date_list[1].zfill(2)
            search_ = '-'.join(date_list)
            break
        else:
            print('Wrong formant.Please re-enter year/month.\n')
            continue

    search_ += '-01'

    wb = openpyxl.load_workbook(target_file)
    result_list = [['ID','TOTAL']]
    for id in ids:
        result_list.append([id, value_get(wb[base_value_sheet], wb[decrement_value_sheet], id, search_)])

    new_wb = openpyxl.Workbook()
    ws = new_wb.active

    for y, row in enumerate(result_list):
        for x, cell in enumerate(row):
            ws.cell(row=y+1,
                    column=x+1,
                    value=result_list[y][x])

    new_wb.save('result.xlsx')

if __name__ == '__main__':
    main()